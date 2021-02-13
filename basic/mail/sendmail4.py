'''
gmail

'''

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
import smtplib, os

# SMTP 접속을 위한 서버, 계정 설정
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
# 보내는 메일 계정
SMTP_USER =  'cbaeck1@gmail.com'# "보내는 사람 메일 주소"
SMTP_PASSWORD = 'oiudntprqfmafdta' # "비밀번호"

# 이메일 유효성 검사 함수
def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
        
# 이메일 보내기 함수
def send_mail(addr, subj_layout, name_layout, rej_layout, imagefiles=None, attachments=None):
    if not is_valid(addr):
        print("Wrong email: " + addr)
        return
    
    # 텍스트 파일
    msg = MIMEMultipart("alternative")

    # 이미지파일이 있는 경우 mixed로 multipart 생성
    if imagefiles:
        msg = MIMEMultipart('mixed')

    # 첨부파일이 있는 경우 mixed로 multipart 생성
    if attachments:
        msg = MIMEMultipart('mixed')


    msg["From"] = SMTP_USER
    msg["To"] = addr
    msg["Subject"] = subj_layout
    # 안녕하세요 xxx 대표님
    text = MIMEText(_text = name_layout, _charset = "utf-8")
    msg.attach(text)


    # 이미지파일이 있으면
    if imagefiles:
        with open(imagefiles, 'rb') as img_file:
            mime_img = MIMEImage(img_file.read())
            # <img src="cid:my_image1">
            # mime_img.add_header('Content-ID', '<' + str_cid_name + '>')
            mime_img.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', imagefiles))

        msg.attach(mime_img)

    # 수신거부문구
    text = MIMEText(_text = rej_layout, _charset = "utf-8")
    msg.attach(text)

    # 첨부파일이 있으면
    if attachments:
        #for attachment in attachments:
        #if attachment:
        file_data = MIMEBase("application", "octect-stream")
        file_data.set_payload(open(attachments, "rb").read())
        encoders.encode_base64(file_data)
        filename = os.path.basename(attachments)
        file_data.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', filename))
        msg.attach(file_data)    

    # smtp로 접속할 서버 정보를 가진 클래스변수 생성
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    # 메일 발송
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    # 닫기
    smtp.close()


from openpyxl import load_workbook
wb = load_workbook('basic/abc.xlsx')
ws = wb.active
subject = '대표님 중대재해법 준비는 하셨습니까? 코로나등의 재해도 든든하게 준비하셔야 합니다.'
reject = """
<div><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">상담을 원하시면&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[상담요청]</b></a><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">를 클릭하세요.&nbsp;</span>If you want to consult, please click the<span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[Request for consultation</b><b>]</b></a></div>
"""
attachment = '기업복지보장_가로.jpg'
for row in ws.iter_rows():
    addr = row[0].value
    name = row[1].value

    #print(addr, subj_layout, rej_layout, attachment)
    send_mail(addr, subject, name, reject, attachment)
