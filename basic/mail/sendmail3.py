'''
gmail

'''

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib, os

# SMTP 접속을 위한 서버, 계정 설정
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
# 보내는 메일 계정
SMTP_USER =  'cbaeck1@gmail.com'# "보내는 사람 메일 주소"
SMTP_PASSWORD = 'oiudntprqfmafdta' # "비밀번호"

# 이메일 유효성 검사 함수
def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
        
# 이메일 보내기 함수
def send_mail(addr, subj_layout, cont_layout, attachments=None):
    if not is_valid(addr):
        print("Wrong email: " + addr)
        return
    
    # 텍스트 파일
    msg = MIMEMultipart("alternative")
    # 첨부파일이 있는 경우 mixed로 multipart 생성
    if attachments:
        msg = MIMEMultipart('mixed')

    msg["From"] = SMTP_USER
    msg["To"] = addr
    msg["Subject"] = subj_layout
    contents = cont_layout
    text = MIMEText(_text = contents, _charset = "utf-8")
    msg.attach(text)
    # 첨부파일이 있으면
    if attachments:
        #for attachment in attachments:
        #if attachment:
        file_data = MIMEBase("application", "octect-stream")
        file_data.set_payload(open(attachment, "rb").read())
        encoders.encode_base64(file_data)
        filename = os.path.basename(attachment)
        file_data.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', filename))
        msg.attach(file_data)    

    # smtp로 접속할 서버 정보를 가진 클래스변수 생성
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    # 메일 발송
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    # 닫기
    smtp.close()


from openpyxl import load_workbook
wb = load_workbook('basic/abc.xlsx')
ws = wb.active
subj_layout = ''
cont_layout = ''
attachment = ''
for i, row in enumerate(ws.iter_rows()):
    addr = row[0].value
    if i == 0:
        subj_layout = row[1].value
        cont_layout = row[2].value
        attachment = row[3].value

    #print(addr, subj_layout, cont_layout, attachment)
    send_mail(addr, subj_layout, cont_layout, attachment)
