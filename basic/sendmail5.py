'''
gmail

e메일에 보낼 컨텐츠를 HTML형태로 생성할 클래스를 만듭니다. (이미지파일 파라미터 있음)


'''
import os, copy
import smtplib               # SMTP 라이브러리
from string import Template  # 문자열 템플릿 모듈
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage


def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False

class EmailHTMLImageContent:
    """e메일에 담길 이미지가 포함된 컨텐츠"""
    def __init__(self, subjectTitle, imageFileName, cidName, template, template_params):
        """이미지파일(str_image_file_name), 컨텐츠ID(str_cid_name)사용된 string template과 딕셔너리형 template_params받아 MIME 메시지를 만든다"""
        assert isinstance(template, Template)
        assert isinstance(template_params, dict)
        self.msg = MIMEMultipart()
        
        # e메일 제목을 설정한다
        self.msg['Subject'] = subjectTitle # e메일 제목을 설정한다
        
        # e메일 본문을 설정한다
        str_msg  = template.safe_substitute(**template_params) # ${변수} 치환하며 문자열 만든다
        mime_msg = MIMEText(str_msg, 'html')                   # MIME HTML 문자열을 만든다
        self.msg.attach(mime_msg)
        
        # e메일 본문에 이미지를 임베딩한다
        assert template.template.find("cid:" + cidName) >= 0, 'template must have cid for embedded image.'
        assert os.path.isfile(imageFileName), 'image file does not exist.'        
        with open(imageFileName, 'rb') as img_file:
            mime_img = MIMEImage(img_file.read())
            mime_img.add_header('Content-ID', '<' + cidName + '>')
        self.msg.attach(mime_img)
        
    def get_message(self, str_to_eamil_addrs):
        """발신자, 수신자리스트를 이용하여 보낼메시지를 만든다 """
        mm = copy.deepcopy(self.msg)
        mm['From'] = SMTP_USER          # 발신자 
        mm['To']   = str_to_eamil_addrs # 수신자리스트 
        return mm


class EmailSender:
    """e메일 발송자"""
    def __init__(self, str_host, num_port=25):
        """호스트와 포트번호로 SMTP로 연결한다 """
        self.str_host = str_host
        self.num_port = num_port
        # self.ss = smtplib.SMTP(host=str_host, port=num_port)
        # SMTP인증이 필요하면 아래 주석을 해제하세요.
        # self.ss.starttls() # TLS(Transport Layer Security) 시작
        # self.ss.login(SMTP_USER, SMTP_PASSWORD) # 메일서버에 연결한 계정과 비밀번호
        self.smtp = smtplib.SMTP_SSL(str_host, num_port)
        self.smtp.login(SMTP_USER, SMTP_PASSWORD)

    def send_message(self, emailContent, str_to_eamil_addrs):
        if not is_valid(str_to_eamil_addrs):
            print("Wrong email: " + str_to_eamil_addrs)
            return

        """e메일을 발송한다 """
        cc = emailContent.get_message(str_to_eamil_addrs)
        self.smtp.send_message(cc, from_addr=SMTP_USER, to_addrs=str_to_eamil_addrs)
        del cc

# SMTP 접속을 위한 서버, 계정 설정
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
# 보내는 메일 계정
SMTP_USER =  'cbaeck1@gmail.com'# "보내는 사람 메일 주소"
SMTP_PASSWORD = 'oiudntprqfmafdta' # "비밀번호"

from openpyxl import load_workbook
wb = load_workbook('basic/abc.xlsx')
ws = wb.active
template = Template("""<html>
                            <head></head>
                            <body>
                                안녕하세요 ${NAME} 대표님.<br><br>
                                <img src="cid:my_image1"><br>
                                <div><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">상담을 원하시면&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[상담요청]</b></a><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">를 클릭하세요.&nbsp;</span>If you want to consult, please click the<span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[Request for consultation</b><b>]</b></a></div>
                                <div><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">메일 수신을 원치 않으시면&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[수신거부]</b></a><span style="color:rgb(153,153,153);font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px">를 클릭하세요. If you don't want this type of information or e-mail, please click the&nbsp;</span><a href="mailto:cbaeck1@gmail.com" rel="noreferrer noopener" style="font-family:&quot;\00b9d1\00c740  \00ace0\00b515&quot;,Helvetica,sans-serif;font-size:12px" target="_blank"><b>[unsubscription]</b></a><div class="yj6qo"></div><div class="adL"><br></div></div>
                            </body>
                        </html>""")

subjectTitle = "대표님 중대재해법 준비는 하셨습니까? 코로나등의 재해도 든든하게 준비하셔야 합니다."
imageFileName = "기업복지보장_가로.jpg"
cidName = 'my_image1'

for i, row in enumerate(ws.iter_rows()):
    addr = row[0].value
    name = row[1].value

    #print(addr, subj_layout, cont_layout, attachment)
    #send_mail(addr, subj_layout, cont_layout, attachment)
    template_params = {'NAME':name}
    emailHTMLImageContent = EmailHTMLImageContent(subjectTitle, imageFileName, cidName, template, template_params)

    str_to_eamil_addrs = addr # 수신자리스트
    emailSender = EmailSender(SMTP_SERVER, SMTP_PORT)
    emailSender.send_message(emailHTMLImageContent, str_to_eamil_addrs)


